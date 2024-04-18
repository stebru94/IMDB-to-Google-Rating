from tqdm import tqdm
import openpyxl
from openpyxl import load_workbook
from configparser import ConfigParser
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import re
import sys
import time

# Load the config.ini file
config = ConfigParser()
config.read('config.ini')

# Get values ​​from sections of the configuration file
excel_file_name = config.get('parameters', 'excel_file_name')
URL_IMDB = config.get('parameters', 'URL_IMDB')
pause_after_cookies = config.getboolean('parameters', 'pause_after_cookies')
pages_to_reveal = config.getint('parameters', 'pages_to_reveal')
suffix = config.get('parameters', 'suffix')


'''
# INPUT
excel_file_name = 'SerieTV_American.xlsx'
URL_IMDB = "https://www.imdb.com/search/title/?title_type=tv_series&release_date=1985-01-01,2023-12-31&countries=US"
pages_to_reveal = 3
'''

MOVIES_TO_FIND = []
CLASSIFIED_MOVIES = []
waste = 0

def pressButtonMore():
    # Find the "50 more" button and the results box
    buttonMore = driver.find_element(By.XPATH, "//span[@class='ipc-see-more__text' and text()='Altri 50']")
    
    # Scroll to the "50 more" button
    driver.execute_script("arguments[0].scrollIntoView();", buttonMore)
    
    # Please wait a short period of time to allow scrolling to complete
    time.sleep(1)
    
    # Get the viewport dimensions
    viewport_height = driver.execute_script("return window.innerHeight;")
    
    # Get the position of the button relative to the beginning of the document
    button_location = buttonMore.location_once_scrolled_into_view
    
    # Calculate the vertical position of the button in the viewport
    button_position_in_viewport = button_location['y']
    
    # Vertical position of the viewport center
    viewport_center = viewport_height / 2
    
    # Calculate the offset needed to center the button on the screen
    scroll_offset = button_position_in_viewport - viewport_center
    
    # Scroll to center the button on the screen
    driver.execute_script("window.scrollBy(0, arguments[0]);", scroll_offset)
    
    # Please wait a short period of time to allow scrolling to complete
    time.sleep(1)
    
    # Click the "50 more" button
    buttonMore.click()
    
    # Please wait a short period of time for the action to complete
    time.sleep(1)



def pause():
    print("press ENTER to continue...")
    input()


def getTitles():
    # Open the web page
    driver.get(URL_IMDB)

    # Wait for the page to load
    driver.implicitly_wait(1)

    # click on I accept all cookies
    try:
        button = driver.find_element(By.XPATH, '//*[@id="__next"]/div/div/div[2]/div/button[2]')
        button.click()
    except:
        pass

    if pause_after_cookies == True:
        pause() 
    
    # Click on "more" multiple times (start, end, step)
    for i in range(1, pages_to_reveal, 1):
        pressButtonMore()
        print("Checking IMDB...", "pag. ",i+1)

    time.sleep(1)
    
    # Find the results container
    content = driver.find_elements(By.CLASS_NAME, 'ipc-title__text')

    # Extract data for each result and format the title
    for title_element in content:
        title = title_element.text
        title_without_number = re.sub(r'^\d+\.\s*', '', title)
        MOVIES_TO_FIND.append(title_without_number)
    '''
    # Print titles added to list
    for title in MOVIES_TO_FIND:
        print(title)
    '''
    print("movie founded: ", len(MOVIES_TO_FIND))


def formatTitle(title):
    # Replace spaces with plus signs (+)
    formatted_title = title.replace(" ", "+")
    formatted_title = formatted_title + " " + suffix
    return formatted_title


def searchMovie(movie, progress_bar):
    # Open the web page
    driver.get("https://www.google.com/search?q=" + formatTitle(movie))

    # Wait for the page to load
    driver.implicitly_wait(1)

    # click on I accept all cookies
    try:
        button = driver.find_element(By.ID, "L2AGLb")
        button.click()
    except:
        pass

    year = None
    genre = None
    duration = None

    # Wait up to 10 seconds for the item to appear with the class ".a19vA"
    try:
        content = WebDriverWait(driver, 1).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".a19vA"))
        )

        # Extract data for each result
        score_text = content.text

        score_match = re.search(r'\d+%', score_text)

        # extracting year, genre e duration
        metadata_movie = driver.find_elements(By.XPATH,
                                             '//*[@id="rcnt"]/div[2]/div/div/div[3]/div/div[1]/div/div/div/div[2]/div[1]/div')
        for (data) in metadata_movie:
            string = data.text

            # Find the year using a regular expression
            year_match = re.search(r'\b\d{4}\b', string)
            year = year_match.group() if year_match else None

            # Find the genre using a regular expression
            genre_match = re.search(r'(?<=‧ ).*?(?= ‧)', string)
            genre = genre_match.group() if genre_match else None

            # Find the duration using a regular expression
            duration_match = re.search(r'\d+h \d+m', string)
            duration = duration_match.group() if duration_match else None

            '''
            # Print extracted data
            print("year:", year)
            print("genre:", genre)
            print("duration:", duration)
            '''
        if score_match:
            score = score_match.group()
            CLASSIFIED_MOVIES.append((movie, score, year, genre, duration))
            progress_bar.update(1)
            #print(f"{contatore_rimanenti} movie rimanenti")
        '''
        else:
            print(f"{movie}: N.D.")
        '''
    except TimeoutException:
        #print(f"{movie}: N.D.")
        pass

# Load the Excel file and read the movies already present
existing_movies = set()
try:
    print(excel_file_name)
    workbook = load_workbook(excel_file_name)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        existing_movies.add(row[0])
except FileNotFoundError:
    pass  # If the file does not exist, there are no movies already present

# Set the driver path
driver_path = "C:/Chromedriver/chromedriver.exe"

# Configure driver options
options = webdriver.ChromeOptions()

# I disable Log messages
options.add_argument("--log-level=3")  

# Initialize the Chrome driver with options
driver = webdriver.Chrome(service=Service(driver_path), options=options)

getTitles()

# Initializing the progress bar
upper_bar_limit = len(MOVIES_TO_FIND)
progress_bar = tqdm(total=upper_bar_limit, desc="Scanning movies..")

for movie in MOVIES_TO_FIND:
    if movie in existing_movies:
        #print(f"The movie '{movie}' is already present in the database. Skipping the Google search.")
        waste = waste+1
        progress_bar.update()

    else:
        searchMovie(movie, progress_bar)

# Close the progress bar
progress_bar.close()

# Sort items by score (in descending order)
CLASSIFIED_MOVIES.sort(key=lambda x: int(x[1].rstrip('%')), reverse=True)

# Print the sorted items
for movie, score, year, genre, duration in CLASSIFIED_MOVIES:
    print(f"{movie}: {score}, {year}, {genre}, {duration}")

# Close the driver
driver.quit()

# Check if the Excel file already exists
try:
    workbook = load_workbook(excel_file_name)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()  # using openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'title'
    sheet['B1'] = 'Percentuale'
    sheet['C1'] = 'year'
    sheet['D1'] = 'genre'
    sheet['E1'] = 'duration'

# Add movie data to the Excel sheet only if it is not already present
for movie, score, year, genre, duration in CLASSIFIED_MOVIES:
    present_movie = False
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if movie in row:
            present_movie = True
            break
    if not present_movie:
        sheet.append([movie, score, year, genre, duration])

# Save the Excel sheet
workbook.save(excel_file_name)
